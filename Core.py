# Imports
from tkinter import Tk, Canvas, PhotoImage, Button, Label, Toplevel, Entry
from time import sleep
from datetime import date
from json import load, dumps
from threading import Thread

# Importing the student list
f1 = open('Assets/Students.txt', 'r')
str1 = f1.read()
stu_list = str1.split(';')
stu_list.sort()
f1.close()

# Pre-Defining some variables to use later
current_stu = 0

more_than_10_entries = False
dock_is_expanded = False
is_topmost = False
done_assigning = False

transparent_color = '#ffffb3'

temp_unassigned_list = list(stu_list)
att_list = []
abs_list = []
exc_list = []
skiped_list = []

att_str = ''
att_str2 = ''

abs_str = ''
abs_str2 = ''

exc_str = ''
exc_str2 = ''

# Starting the main window and setting some parameters
root = Tk()
root.geometry('300x110')
root.title('Build 1.3.2')
root.resizable(False, False)
root.overrideredirect(1)
root.attributes('-topmost', True)
root.wm_attributes('-transparentcolor', transparent_color)

# Loading The Assets
bg_Image = PhotoImage(file='Assets/1.png')
yes_button_unpres = PhotoImage(file='Assets/Button_Yes_Unpressed.png')
yes_button_pres = PhotoImage(file='Assets/Button_Yes_Pressed.png')
no_button_unpres = PhotoImage(file='Assets/Button_No_Unpressed.png')
no_button_pres = PhotoImage(file='Assets/Button_No_Pressed.png')
excused_button_unpres = PhotoImage(file='Assets/Button_Excused_Unpressed.png')
excused_button_pres = PhotoImage(file='Assets/Button_Excused_Pressed.png')
skip_button_unpres = PhotoImage(file='Assets/Button_Skip_Unpressed.png')
skip_button_pres = PhotoImage(file='Assets/Button_Skip_Pressed.png')

dock_but = PhotoImage(file='Assets/Button_Dock.png')
dock_green_image = PhotoImage(file='Assets/2_1.png')
dock_red_image = PhotoImage(file='Assets/3_1.png')
dock_yellow_image = PhotoImage(file='Assets/4_1.png')
dock_green_list = PhotoImage(file='Assets/Button_Greenlist.png')
dock_red_list = PhotoImage(file='Assets/Button_Redlist.png')
dock_yellow_list = PhotoImage(file='Assets/Button_Yellowlist.png')

reassign_button_img = PhotoImage(file='Assets/Button_Reassign.png')
reassign_window_bg = PhotoImage(file='Assets/Reassign_bg.png')
reassign_button_att = PhotoImage(file='Assets/reassign_attending_button.png')
reassign_button_abs = PhotoImage(file='Assets/reassign_absent_button.png')
reassign_button_exc = PhotoImage(file='Assets/reassign_excused_button.png')

x_button_img = PhotoImage(file='Assets/X_button.png')
topmost_button_img = PhotoImage(file='Assets/topmost_button.png')
line = PhotoImage(file='Assets/300x2_line.png')

# Function: Compiling the Attending List String
def att_str_com():
    global att_str
    global att_str2
    global more_than_10_entries
    more_than_10_entries = False

    returned_strs = [None, None]

    att_str = ''
    att_str2 = ''

    var1 = 0
    for x in att_list:
        if var1 == 0:
            att_str = str(var1+1) + '  |  ' + str(att_list[var1]).capitalize()
            var1 += 1

        elif var1 == 12:
            more_than_10_entries = True
            att_str2 = str(var1+1) + '  |  ' + str(att_list[var1]).capitalize()
            var1 += 1

        else:
            if var1 >= 12:
                more_than_10_entries = True
                # print('a')
                att_str2 = att_str2 + '\n\n' + str(var1+1) + '  |  ' + str(att_list[var1]).capitalize()
                # print(att_str)
                var1 += 1
                # print(var1)

            else:
                att_str = att_str + '\n\n' + str(var1+1) + '  |  ' + str(att_list[var1]).capitalize()
                # print(att_str)
                var1 += 1
                # print(var1)

    if more_than_10_entries:
        returned_strs = [att_str, att_str2]

    else:
        returned_strs = [att_str, None]

    return returned_strs

# Function: Compiling the Absent List String
def abs_str_com():
    global abs_str
    global abs_str2
    global more_than_10_entries
    more_than_10_entries = False

    returned_strs = [None, None]

    abs_str = ''
    abs_str2 = ''

    var1 = 0
    for x in abs_list:
        if var1 == 0:
            abs_str = str(var1+1) + '  |  ' + str(abs_list[var1]).capitalize()
            var1 += 1

        elif var1 == 12:
            abs_str2 = str(var1+1) + '  |  ' + str(abs_list[var1]).capitalize()
            var1 += 1

        else:
            if var1 >= 12:
                more_than_10_entries = True
                # print('a')
                abs_str2 = abs_str2 + '\n\n' + str(var1+1) + '  |  ' + str(abs_list[var1]).capitalize()
                # print(abs_str)
                var1 += 1
                # print(var1)

            else:
                abs_str = abs_str + '\n\n' + str(var1+1) + '  |  ' + str(abs_list[var1]).capitalize()
                # print(abs_str)
                var1 += 1
                # print(var1)

    if more_than_10_entries:
        returned_strs = [abs_str, abs_str2]

    else:
        returned_strs = [abs_str, None]

    return returned_strs

# Function: Compiling the Excused List String
def exc_str_com():
    global exc_str
    global exc_str2
    global more_than_10_entries
    more_than_10_entries = False

    returned_strs = [None, None]

    exc_str = ''
    exc_str2 = ''

    var1 = 0
    for x in exc_list:
        if var1 == 0:
            exc_str = str(var1+1) + '  |  ' + str(exc_list[var1]).capitalize()
            var1 += 1

        elif var1 == 12:
            exc_str2 = str(var1+1) + '  |  ' + str(exc_list[var1]).capitalize()
            var1 += 1

        else:
            if var1 >= 12:
                more_than_10_entries = True
                # print('a')
                exc_str2 = exc_str2 + '\n\n' + str(var1+1) + '  |  ' + str(exc_list[var1]).capitalize()
                # print(exc_str)
                var1 += 1
                # print(var1)

            else:
                exc_str = exc_str + '\n\n' + str(var1+1) + '  |  ' + str(exc_list[var1]).capitalize()
                # print(exc_str)
                var1 += 1
                # print(var1)

    if more_than_10_entries:
        returned_strs = [exc_str, exc_str2]

    else:
        returned_strs = [exc_str, None]

    return returned_strs

# Function: Expands the Lists Dock
def dock_expand():
    global c2
    global dock_green_list_button
    global dock_red_list_button
    global dock_yellow_list_button
    global b_dock_2
    global stu_list_lbl
    global dock_is_expanded

    dock_is_expanded = True

    b_dock.place_forget()

    root.geometry('300x605')

    c2 = Canvas(root, width=300, height=612, bd=-2, bg=transparent_color)
    c2.create_image(0, 0, image=dock_green_image, anchor='nw')
    c2.place(x=0, y=100, anchor='nw')

    dock_green_list_button = Button(root, image=dock_green_list, relief='flat', bd=-2, command=_show_att_stus, bg='black', activebackground='black')
    dock_green_list_button.place(x=50, y=583, anchor='center')

    dock_red_list_button = Button(root, image=dock_red_list, relief='flat', bd=-2, command=_show_abs_stus, bg='Black', activebackground='black')
    dock_red_list_button.place(x=150, y=583, anchor='center')

    dock_yellow_list_button = Button(root, image=dock_yellow_list, relief='flat', bd=-2, command=_show_exc_stus, bg='black', activebackground='black')
    dock_yellow_list_button.place(x=250, y=583, anchor='center')

    b_dock_2 = Button(root, image=dock_but, bd=0, bg=transparent_color, activebackground=transparent_color, command=dock_shrink)
    b_dock_2.place(x=150, y=598, anchor='center')

    _show_att_stus(first_boot=True)

# Function: Shrinks (destroy) the Lists Dock
def dock_shrink():
    global dock_is_expanded

    dock_is_expanded = False

    root.geometry('300x110')
    c2.destroy()
    l2.destroy()
    if more_than_10_entries:
        l3.destroy()

    b_dock.place(x=150, y=106, anchor='center')

# Function: Post Assignment Cleanup
def all_assigned():
    global done_assigning
    done_assigning = True

    l1.place_forget()

    b1.destroy()
    b2.destroy()
    b3.destroy()
    b4.destroy()

    l1.configure(text='All students assigned')
    l1.place(x=150, y=25, anchor='center')

    reassign_button.place(x=150, y=75, anchor='center')

# Function: Resets the Assignment Panel every assignment
def panel_Reset():
    global bg_Image

    bg_Image = PhotoImage(file='Assets/1.png')
    c1.create_image(0, 0, image=bg_Image, anchor='nw')

    b1.place_forget()
    b2.place_forget()
    b3.place_forget()
    b4.place_forget()

    b1.configure(image=yes_button_unpres, state='active')
    b2.configure(image=no_button_unpres, state='active')
    b3.configure(image=excused_button_unpres, state='active')
    b4.configure(image=skip_button_unpres, state='active')

    b1.place(x=200, y=55, anchor='center')
    b2.place(x=250, y=80, anchor='center')
    b3.place(x=50, y=80, anchor='center')
    b4.place(x=100, y=55, anchor='center')

# Function: Progresses Through the Student List
def next_stu(another_skip=None):
    global current_stu
    global temp_unassigned_list
    global skiped_list
#    current_stu += 1
    print(current_stu)
    panel_Reset()

    try:
        l1.configure(text=f'Is {temp_unassigned_list[current_stu].capitalize()} Here ?')

    except IndexError:
        if skiped_list:
            if not temp_unassigned_list:
                temp_unassigned_list = list(skiped_list)
                skiped_list = []
                current_stu = 0

            l1.configure(text=f'Is {temp_unassigned_list[current_stu].capitalize()} Here ?')

        elif not skiped_list:
            all_assigned()

        else:
            print('List Error')

    except Exception as e:
        print(e)
        input()
        exit()

# Function: Reassigning Students
def reassign_stu():
    global stage
    global origin
    global index
    global destination

    origin = None
    index = 0
    destination = None
    stage = 1

    window2 = Toplevel()
    window2.geometry(f'{300}x{250}+{300}+{0}')
    window2.overrideredirect(1)
    window2.resizable(False, False)
    window2.attributes('-topmost', True)
    window2.wm_attributes('-transparentcolor', transparent_color)

    c2 = Canvas(window2, width=300, height=250, bd=-2, bg=transparent_color)
    c2.create_image(0, 0, image=reassign_window_bg, anchor='nw')
    c2.place(x=0, y=0, anchor='nw')

    def stage_switch(data=None):
        global stage
        global origin
        global index
        global destination

        if stage == 1:
            origin = data
            stage += 1
            stage_2()

        elif stage == 2:
            index = data
            stage += 1
            stage_3()

        elif stage == 3:
            destination = data
            reassign_func()

    def stage_1():
        global l4
        global rb1
        global rb2
        global rb3

        l4 = Label(window2, text='What list is the Student currently in?', bg='#323232', fg='#ffffff', font=('Micross', 10))
        l4.place(x=150, y=16, anchor='center')

        rb1 = Button(window2, image=reassign_button_att, bd=0, relief='flat', bg='#007f00', activebackground='#007f00', command=lambda: stage_switch('att'))
        rb1.place(x=150, y=70, anchor='center')
        rb2 = Button(window2, image=reassign_button_abs, bd=0, relief='flat', bg='#7f0000', activebackground='#7f0000', command=lambda: stage_switch('abs'))
        rb2.place(x=245, y=70, anchor='center')
        rb3 = Button(window2, image=reassign_button_exc, bd=0, relief='flat', bg='#7f7f00', activebackground='#7f7f00', command=lambda: stage_switch('exc'))
        rb3.place(x=55, y=70, anchor='center')


    def stage_2():
        global l5
        global e1

        rb1.configure(state='disabled')
        rb2.configure(state='disabled')
        rb3.configure(state='disabled')

        c2.create_image(150, 110, image=line, anchor='center')

        l5 = Label(window2, text='What is the Index Number of the Student? ', bg='#323232', fg='#ffffff', font=('Micross', 10))
        l5.place(x=10, y=130, anchor='w')

        e1 = Entry(window2, bd=0, relief='flat', width=2)
        e1.place(x=280, y=130, anchor='e')

        def theard_func():
            while True:
                if e1.get() != '':
                    stage_switch(int(e1.get()))
                    break

        T1 = Thread(target=theard_func)
        T1.start()


    def stage_3():
        global l6
        global rb4
        global rb5
        global rb6

        e1.configure(state='disabled')

        c2.create_image(150, 150, image=line, anchor='center')

        l6 = Label(window2, text='Move the Student to What List?', bd=0, fg='#ffffff', bg='#323232', font=('micross', 10))
        l6.place(x=150, y=170, anchor='center')

        rb4 = Button(window2, image=reassign_button_att, bd=0, relief='flat', bg='#007f00', activebackground='#007f00', command=lambda: stage_switch('att'))
        rb4.place(x=150, y=215, anchor='center')
        rb5 = Button(window2, image=reassign_button_abs, bd=0, relief='flat', bg='#7f0000', activebackground='#7f0000', command=lambda: stage_switch('abs'))
        rb5.place(x=245, y=215, anchor='center')
        rb6 = Button(window2, image=reassign_button_exc, bd=0, relief='flat', bg='#7f7f00', activebackground='#7f7f00', command=lambda: stage_switch('exc'))
        rb6.place(x=55, y=215, anchor='center')

    def reassign_func():
        global att_list
        global abs_list
        global exc_list
        global origin
        global index
        global destination

        index -= 1

        if origin == 'att':
            if destination == 'att':
                if dock_is_expanded:
                    _show_att_stus()

            elif destination == 'abs':
                abs_list.append(att_list[index])
                att_list.pop(index)
                if dock_is_expanded:
                    _show_abs_stus()

            else:
                exc_list.append(att_list[index])
                att_list.pop(index)
                if dock_is_expanded:
                    _show_exc_stus()

        elif origin == 'abs':
            if destination == 'att':
                att_list.append(abs_list[index])
                abs_list.pop(index)
                if dock_is_expanded:
                    _show_att_stus()

            elif destination == 'abs':
                if dock_is_expanded:
                    _show_abs_stus()

            else:
                exc_list.append(abs_list[index])
                abs_list.pop(index)
                if dock_is_expanded:
                    _show_exc_stus()

        else:
            if destination == 'att':
                att_list.append(exc_list[index])
                exc_list.pop(index)
                if dock_is_expanded:
                    _show_att_stus()

            elif destination == 'abs':
                abs_list.append(exc_list[index])
                exc_list.pop(index)
                if dock_is_expanded:
                    _show_abs_stus()

            else:
                if dock_is_expanded:
                    _show_exc_stus()

        cancel_reassign()

    def cancel_reassign():
        try:
            window2.destroy()
            c2.destroy()
            l4.destroy()
            l5.destroy()
            l6.destroy()
            rb1.destroy()
            rb2.destroy()
            rb3.destroy()
            rb4.destroy()
            rb5.destroy()
            rb6.destroy()
            e1.destroy()
            cancel_button.destroy()
        except Exception:
            pass

    cancel_button = Button(window2, image=x_button_img, relief='flat', bd=0, bg='#323232', activebackground='#323232', command=cancel_reassign)
    cancel_button.place(x=300, y=0, anchor='ne')

    stage_1()

# Function: Handel when Attending Button is pressed
def yes_pressed():
    global bg_Image
    global current_stu

    att_list.append(temp_unassigned_list[current_stu])
    temp_unassigned_list.pop(current_stu)
    att_list.sort()

    b2.place_forget()
    b3.place_forget()
    b4.place_forget()

    b1.configure(image=yes_button_pres, state='disabled')
    bg_Image = PhotoImage(file='Assets/2.png')
    c1.create_image(0, 0, image=bg_Image, anchor='nw')
    root.update()
    sleep(.5)
    next_stu()

    if dock_is_expanded:
        _show_att_stus()

# Function: Handel when Absent Button is pressed
def no_pressed():
    global bg_Image
    global current_stu

    abs_list.append(temp_unassigned_list[current_stu])
    temp_unassigned_list.pop(current_stu)
    abs_list.sort()

    b1.place_forget()
    b3.place_forget()
    b4.place_forget()

    b2.configure(image=no_button_pres, state='disabled')
    bg_Image = PhotoImage(file='Assets/3.png')
    c1.create_image(0, 0, image=bg_Image, anchor='nw')
    root.update()
    sleep(.5)
    next_stu()

    if dock_is_expanded:
        _show_abs_stus()

# Function: Handel when Excused Button is pressed
def excused_pressed():
    global bg_Image
    global current_stu

    exc_list.append(temp_unassigned_list[current_stu])
    temp_unassigned_list.pop(current_stu)
    exc_list.sort()

    b1.place_forget()
    b2.place_forget()
    b4.place_forget()

    b3.configure(image=excused_button_pres, state='disabled')
    bg_Image = PhotoImage(file='Assets/4.png')
    c1.create_image(0, 0, image=bg_Image, anchor='nw')
    root.update()
    sleep(1)
    next_stu()

    if dock_is_expanded:
        _show_exc_stus()

# Function: Handel when Skip Button is pressed
def _skip_pressed():
    global bg_Image
    global current_stu

    b1.place_forget()
    b2.place_forget()
    b3.place_forget()

    if temp_unassigned_list:
        skiped_list.append(temp_unassigned_list[current_stu])
        temp_unassigned_list.pop(current_stu)

    else:
        pass

    b4.configure(image=skip_button_pres, state='disabled')
    bg_Image = PhotoImage(file='Assets/5.png')
    c1.create_image(0, 0, image=bg_Image, anchor='nw')
    root.update()
    sleep(.5)
    next_stu(True)

# Function: Switch to Attending list in the List Dock
def _show_att_stus(first_boot=False):
    global l2
    global l3

    if more_than_10_entries:
        l3.destroy()

    c2.create_image(0, 0, image=dock_green_image, anchor='nw')

    att_stu_str = att_str_com()

    if first_boot:
        l2 = Label(root, text=att_stu_str[0], font=('micross', 12), bg='#323232', fg='white', justify='left')
        l2.place(x=16, y=120, anchor='nw')

        if more_than_10_entries:
            l3 = Label(root, text=att_stu_str[1], font=('micross', 12), bg='#323232', fg='white', justify='left')
            l3.place(x=150, y=120, anchor='nw')

    else:
        l2.configure(text=att_stu_str[0])
        if more_than_10_entries:
            try:
                l3.destroy()

            except NameError:
                pass

            l3 = Label(root, text=att_stu_str[1], font=('micross', 12), bg='#323232', fg='white', justify='left')
            l3.place(x=150, y=120, anchor='nw')

# Function: Switch to Absent list in the List Dock
def _show_abs_stus():
    global l2
    global l3


    if more_than_10_entries:
        l3.destroy()


    c2.create_image(0, 0, image=dock_red_image, anchor='nw')

    abs_stu_str = abs_str_com()

    l2.configure(text=abs_stu_str[0])

    if more_than_10_entries:
        try:
            l3.destroy()

        except NameError:
            pass

        l3 = Label(root, text=abs_stu_str[1], font=('micross', 12), bg='#323232', fg='white', justify='left')
        l3.place(x=150, y=120, anchor='nw')

# Function: Switch to Excused list in the List Dock
def _show_exc_stus():
    global l2
    global l3

    if more_than_10_entries:
        l3.destroy()

    c2.create_image(0, 0, image=dock_yellow_image, anchor='nw')

    exc_stu_str = exc_str_com()

    l2.configure(text=exc_stu_str[0])

    if more_than_10_entries:
        try:
            l3.destroy()

        except NameError:
            pass

        l3 = Label(root, text=exc_stu_str[1], font=('micross', 12), bg='#323232', fg='white', justify='left')
        l3.place(x=150, y=120, anchor='nw')


# noinspection PyBroadException
# Program Pre-Launch Configurations
def _prog_config(topmost_toggle=None):
    global is_topmost

    if topmost_toggle is not None:
        if not is_topmost:
            is_topmost = True

        elif is_topmost:
            is_topmost = False

        root.wm_attributes('-topmost', is_topmost)
        root.overrideredirect(is_topmost)

        try:
            dock_shrink()

        except Exception:
            pass

        root.geometry('300x110+0+0')
        root.iconbitmap('Assets/icon.ico')

# Function: Handels when the Exit Button is pressed
def _quit_function():
    # for some reason when converted to exe the program do not recognize the command Quit() within the button
    # NameError: name 'Quit' is not defined
    root.quit()
    # print('Bye')
    # quit()


# Creating and placing widgets
c1 = Canvas(root, width=300, height=112, bd=-2)
c1.create_image(0, 0, image=bg_Image, anchor='nw')
c1.place(x=0, y=0, anchor='nw')

b_dock = Button(root, image=dock_but, bd=0, bg=transparent_color, activebackground=transparent_color, command=dock_expand)
b_dock.place(x=150, y=106, anchor='center')

l1 = Label(root, text=f'Is {temp_unassigned_list[current_stu].capitalize()} Here ?', bg='#323232', fg='white', font=('consolas', 12))
l1.place(x=150, y=25, anchor='center')

b1 = Button(root, image=yes_button_unpres, relief='flat', bd=0, command=yes_pressed)
b1.place(x=200, y=55, anchor='center')

b2 = Button(root, image=no_button_unpres, relief='flat', bd=0, command=no_pressed)
b2.place(x=250, y=80, anchor='center')

b3 = Button(root, image=excused_button_unpres, relief='flat', bd=0, command=excused_pressed)
b3.place(x=50, y=80, anchor='center')

b4 = Button(root, image=skip_button_unpres, relief='flat', bd=0, command=_skip_pressed)
b4.place(x=100, y=55, anchor='center')

reassign_button = Button(root, image=reassign_button_img, relief='flat', bd=0, command=reassign_stu)

x_button = Button(root, image=x_button_img, relief='flat', bd=0, bg='#323232', activebackground='#323232', command=_quit_function)
x_button.place(x=0, y=0, anchor='nw')

topmost_button = Button(root, image=topmost_button_img, relief='flat', bd=0, bg='#323232', activebackground='#323232', command=lambda: _prog_config(topmost_toggle=True))
topmost_button.place(x=18, y=0, anchor='nw')

# Compiles the XLSX Excel File
def _compile_xlsx():
    current_month_str = None

    while True:
        if date.today().month == 1:
            current_month_str = "January"
            break

        elif date.today().month == 2:
            current_month_str = "February"
            break

        elif date.today().month == 3:
            current_month_str = "March"
            break

        elif date.today().month == 4:
            current_month_str = "April"
            break

        elif date.today().month == 5:
            current_month_str = "May"
            break

        elif date.today().month == 6:
            current_month_str = "June"
            break

        elif date.today().month == 7:
            current_month_str = "July"
            break

        elif date.today().month == 8:
            current_month_str = "August"
            break

        elif date.today().month == 9:
            current_month_str = "September"
            break

        elif date.today().month == 10:
            current_month_str = "October"
            break

        elif date.today().month == 11:
            current_month_str = "November"
            break

        elif date.today().month == 12:
            current_month_str = "December"
            break

        else:
            input('Systemic Error')
            quit()


    from openpyxl import Workbook
    from openpyxl.styles import Alignment, PatternFill, Border, Side

    from os import replace, getlogin, path

    wb = Workbook()
    ws = wb.active

    thin_border = Side(border_style='thin', color='000000')
    double_border = Side(border_style='double', color='000000')

    ws.merge_cells("A1:D1")

    ws['A1'].value = f'Attendance For {current_month_str}'
    ws['A1'].alignment = Alignment('center')
    ws['A2'] = 'Days/Stus'
    ws['A2'].border = Border(top=double_border, left=double_border, right=double_border, bottom=double_border)
    ws['A2'].alignment = Alignment('center')

    for x in range(31):
        x += 1
        ws[f'A{x + 2}'] = x
        ws[f'A{x + 2}'].alignment = Alignment('center')

    for x in range(len(stu_list)):
        ws[f'{chr(66 + x)}2'] = stu_list[x]
        ws[f'{chr(66 + x)}2'].alignment = Alignment('center')

    for stus in range(len(stu_list)):
        for days in range(31):
            if stu_list[stus] in loaded_json["Attendance_Data"]["Attending"][f'{days + 1}']:
                ws[f'{chr(66 + stus)}{days + 3}'] = 'Att'
                ws[f'{chr(66 + stus)}{days + 3}'].alignment = Alignment('center')
                ws[f'{chr(66 + stus)}{days + 3}'].fill = PatternFill("solid", fgColor='00ff00')
                ws[f'{chr(66 + stus)}{days + 3}'].border = Border(top=thin_border, left=thin_border, right=thin_border, bottom=thin_border)

            elif stu_list[stus] in loaded_json["Attendance_Data"]["Absent"][f'{days + 1}']:
                ws[f'{chr(66 + stus)}{days + 3}'] = 'Abs'
                ws[f'{chr(66 + stus)}{days + 3}'].alignment = Alignment('center')
                ws[f'{chr(66 + stus)}{days + 3}'].fill = PatternFill("solid", fgColor='ff0000')
                ws[f'{chr(66 + stus)}{days + 3}'].border = Border(top=thin_border, left=thin_border, right=thin_border, bottom=thin_border)

            elif stu_list[stus] in loaded_json["Attendance_Data"]["Excused"][f'{days + 1}']:
                ws[f'{chr(66 + stus)}{days + 3}'] = 'Exc'
                ws[f'{chr(66 + stus)}{days + 3}'].alignment = Alignment('center')
                ws[f'{chr(66 + stus)}{days + 3}'].fill = PatternFill("solid", fgColor='ffff00')
                ws[f'{chr(66 + stus)}{days + 3}'].border = Border(top=thin_border, left=thin_border, right=thin_border, bottom=thin_border)

            else:
                ws[f'{chr(66 + stus)}{days + 3}'] = ''
                ws[f'{chr(66 + stus)}{days + 3}'].alignment = Alignment('center')
                ws[f'{chr(66 + stus)}{days + 3}'].border = Border(top=thin_border, left=thin_border, right=thin_border, bottom=thin_border)

    wb.save(f'{current_month_str} Attendance.xlsx')

    for x in range(26):
        if path.exists(f'{chr(65 + x)}:\\Users\\{getlogin()}\\Desktop'):
            replace(f'{current_month_str} Attendance.xlsx', f'{chr(65+x)}:\\Users\\{getlogin()}\\Desktop\\{current_month_str} Attendance.xlsx')
            break

# Pre-Launch Checks
def _pre_execution():
    global loaded_json

    try:
        loaded_json = load(open('Assets/Data.json', 'r'))
        first_boot = False

    except Exception:
        f1 = open('Assets/Data.json', 'w')
        f1.write(f'''{'{'}
          "Settings": {'{'}
            "Last_Export": {date.today().month}
          {'}'},
          "Attendance_Data": {'{'}
            "Attending": {'{'}
              "1": [],
              "2":  [],
              "3":  [],
              "4":  [],
              "5":  [],
              "6":  [],
              "7":  [],
              "8":  [],
              "9":  [],
              "10":  [],
              "11":  [],
              "12":  [],
              "13":  [],
              "14":  [],
              "15":  [],
              "16":  [],
              "17":  [],
              "18":  [],
              "19":  [],
              "20":  [],
              "21":  [],
              "22":  [],
              "23":  [],
              "24":  [],
              "25":  [],
              "26":  [],
              "27":  [],
              "28":  [],
              "29":  [],
              "30":  [],
              "31":  []
            {'}'},
            "Absent": {'{'}
              "1": [],
              "2":  [],
              "3":  [],
              "4":  [],
              "5":  [],
              "6":  [],
              "7":  [],
              "8":  [],
              "9":  [],
              "10":  [],
              "11":  [],
              "12":  [],
              "13":  [],
              "14":  [],
              "15":  [],
              "16":  [],
              "17":  [],
              "18":  [],
              "19":  [],
              "20":  [],
              "21":  [],
              "22":  [],
              "23":  [],
              "24":  [],
              "25":  [],
              "26":  [],
              "27":  [],
              "28":  [],
              "29":  [],
              "30":  [],
              "31":  []
            {'}'},
            "Excused": {'{'}
              "1": [],
              "2":  [],
              "3":  [],
              "4":  [],
              "5":  [],
              "6":  [],
              "7":  [],
              "8":  [],
              "9":  [],
              "10":  [],
              "11":  [],
              "12":  [],
              "13":  [],
              "14":  [],
              "15":  [],
              "16":  [],
              "17":  [],
              "18":  [],
              "19":  [],
              "20":  [],
              "21":  [],
              "22":  [],
              "23":  [],
              "24":  [],
              "25":  [],
              "26":  [],
              "27":  [],
              "28":  [],
              "29":  [],
              "30":  [],
              "31":  []
            {'}'}
          {'}'}
        {'}'}''')
        f1.close()

        loaded_json = load(open('Assets/Data.json', 'r'))
        first_boot = True


    if loaded_json["Settings"]["Last_Export"] != (date.today().month - 1) and loaded_json["Settings"]["Last_Export"] != date.today().month:
        loaded_json["Settings"]["Last_Export"] = date.today().month
        _json_clear()
        Json_Error = True

    else:
        Json_Error = False

    if not first_boot and not Json_Error:
        if 3 >= date.today().day >= 1 and (date.today().month - 1) == loaded_json["Settings"]["Last_Export"]:
            _compile_xlsx()
            _json_clear()

# Clears the Loaded Json File
def _json_clear():
    global loaded_json

    loaded_json = load(open('Assets/Data.json', 'r'))
    for x in range(len(loaded_json["Attendance_Data"]["Attending"])):
        loaded_json["Attendance_Data"]["Attending"][f"{x+1}"] = []

    for x in range(len(loaded_json["Attendance_Data"]["Absent"])):
        loaded_json["Attendance_Data"]["Absent"][f"{x+1}"] = []

    for x in range(len(loaded_json["Attendance_Data"]["Excused"])):
        loaded_json["Attendance_Data"]["Excused"][f"{x+1}"] = []

    loaded_json["Settings"]["Last_Export"] = date.today().month

    f1 = open('Assets/Data.json', 'w')
    f1.write(dumps(loaded_json))
    f1.close()

# Post-Execution Tasks
def _post_completion():
    if done_assigning:
        loaded_json["Attendance_Data"]["Attending"][f"{date.today().day}"] = att_list
        loaded_json["Attendance_Data"]["Absent"][f"{date.today().day}"] = abs_list
        loaded_json["Attendance_Data"]["Excused"][f"{date.today().day}"] = exc_list
        f1 = open('Assets/Data.json', 'w')
        f1.write(dumps(loaded_json))
        f1.close()


root.wm_protocol('WM_DELETE_WINDOW', _quit_function)
# Program Execution Sequence
if __name__ == '__main__':
    _pre_execution()
    root.mainloop()
    _post_completion()
    print(att_list)
    print(abs_list)
    print(exc_list)
    print(skiped_list)
